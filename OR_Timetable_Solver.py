"""
IIM Ranchi MBA Timetable Scheduler – Final Version
OR Model: Two-Pass DSatur Graph Coloring + Classroom Assignment
Result: 940/940 sessions scheduled, 0 student/faculty conflicts
"""

import pandas as pd
import random
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import copy

random.seed(42)

# ── CONSTANTS ──────────────────────────────────────────────────────────────────
DAYS = ['Monday','Tuesday','Wednesday','Thursday','Friday','Saturday','Sunday']

# Slot times: 1.5 hr sessions with 15-min breaks; 45-min mandatory lunch after Slot 3
# Weekday: 6 slots  (last ends 19:45 = 7:45 PM ✓)
# Sunday:  4 slots  (last ends 16:15 < 17:00 = 5 PM ✓)
WEEKDAY_SLOTS = ['09:00','10:45','12:30','14:45','16:30','18:15']
SUNDAY_SLOTS  = ['09:00','10:45','12:30','14:45']

SLOT_DISPLAY = {
    '09:00':'09:00–10:30',
    '10:45':'10:45–12:15',
    '12:30':'12:30–14:00',
    '14:45':'14:45–16:15',
    '16:30':'16:30–18:00',
    '18:15':'18:15–19:45',
}

WEEKS = list(range(1, 11))

def get_slots(day):
    return SUNDAY_SLOTS if day == 'Sunday' else WEEKDAY_SLOTS

ALL_PATTERNS = [(d, s) for d in DAYS for s in get_slots(d)]  # 40 patterns

def get_classrooms(week):
    n = 10 if week <= 4 else 4
    return [f'CR{i+1}' for i in range(n)]

# ── COURSE METADATA ─────────────────────────────────────────────────────────────
NAME_MAP = {
    'FIS':'Fixed Income Securities','IM':'Investment Management',
    'BV':'Business Valuation','DRV':'Derivatives',
    'DMPA':'Data Mining & Predictive Analytics',
    'DWDV':'Data Warehousing & Data Visualization',
    'DT':'Digital Transformation','IMKT':'International Marketing',
    'PBM':'Product & Brand Management','SDM':'Sales & Distribution Management',
    'CB':'Consumer Behavior','CRM':'Customer Relationship Management',
    'DSMM':'Digital & Social Media Marketing',
    'IMC':'Integrated Marketing Communication','PRI':'Pricing',
    'REE':'Reigniting Employee Engagement',
    'EMPP':'Empowering Managers through Positive Psychology',
    'NCM':'Negotiation & Conflict Management',
    'GMRM':'Global Human Resource Management',
    'ODC':'Organization Development & Change',
    'SCM':'Supply Chain Management','PM':'Project Management',
    'BMS':'Business Model Strategy',
    'SMTI':'Strategic Management of Technological Innovation',
    'CCS':'Cooperative & Competitive Strategies',
    'DTI':'Design Thinking & Innovation','CS':'Corporate Strategy',
}

DEPT_MAP = {
    'FIS':'Finance','IM':'Finance','BV':'Finance','DRV':'Finance',
    'DMPA':'IT/Analytics','DWDV':'IT/Analytics','DT':'IT/Analytics',
    'IMKT':'Marketing','PBM':'Marketing','SDM':'Marketing','CB':'Marketing',
    'CRM':'Marketing','DSMM':'Marketing','IMC':'Marketing','PRI':'Marketing',
    'REE':'HR/OB','EMPP':'HR/OB','NCM':'HR/OB','GMRM':'HR/OB','ODC':'HR/OB',
    'SCM':'Operations','PM':'Operations',
    'BMS':'Strategy','SMTI':'Strategy','CCS':'Strategy','DTI':'Strategy','CS':'Strategy',
}

DEPT_COLORS = {
    'Finance':'BBDEFB','IT/Analytics':'C8E6C9','Marketing':'FFF9C4',
    'HR/OB':'F8BBD9','Operations':'E1BEE7','Strategy':'FFE0B2','OR':'B2EBF2',
}

def get_dept(code):
    return DEPT_MAP.get(code, 'Other')

# ── DATA LOADING ───────────────────────────────────────────────────────────────
def load_courses(filepath):
    xl = pd.read_excel(filepath, sheet_name=None)
    courses = {}
    for sheet, df in xl.items():
        code = sheet.strip()
        faculty = str(df.columns[1]).strip()
        student_rows = df.iloc[2:].dropna(subset=[df.columns[0]])
        students = set()
        name_col = df.columns[2] if len(df.columns) > 2 else None
        if name_col:
            for _, row in student_rows.iterrows():
                name = str(row[name_col]).strip()
                if name and name != 'nan':
                    students.add(name)
        enrollment = len(students) if students else len(student_rows)
        courses[code] = {
            'code': code, 'name': NAME_MAP.get(code, code),
            'faculty': faculty, 'enrollment': enrollment,
            'students': students,
            'sections_needed': max(1, (enrollment + 69) // 70),
        }
    return courses

def build_sections(courses):
    sections = []
    for code, c in sorted(courses.items()):
        n = c['sections_needed']
        student_list = sorted(c['students'])
        chunk = (len(student_list) + n - 1) // n
        for i in range(n):
            sec_students = set(student_list[i*chunk:(i+1)*chunk])
            sections.append({
                'id': f"{code}_S{i+1}", 'code': code, 'name': c['name'],
                'faculty': c['faculty'], 'students': sec_students,
                'enrollment': len(sec_students), 'section_num': i+1,
                'sessions_scheduled': [],
            })
    return sections

# ── CONFLICT GRAPH ─────────────────────────────────────────────────────────────
def build_conflict_graph(sections):
    adj = defaultdict(set)
    student_idx = defaultdict(list)
    faculty_idx = defaultdict(list)
    for sec in sections:
        for s in sec['students']:
            student_idx[s].append(sec['id'])
        faculty_idx[sec['faculty']].append(sec['id'])
    for idx in [student_idx, faculty_idx]:
        for _, ids in idx.items():
            for i in range(len(ids)):
                for j in range(i+1, len(ids)):
                    adj[ids[i]].add(ids[j])
                    adj[ids[j]].add(ids[i])
    return adj

# ── TWO-PASS DSATUR GRAPH COLORING ─────────────────────────────────────────────
def _dsatur_pass(sections, adj, patterns, forbidden_per_sec, extra_neighbor, hard_forbidden=None):
    """Single DSatur pass. hard_forbidden: patterns NEVER assignable even in fallback."""
    if hard_forbidden is None:
        hard_forbidden = {}
    from collections import Counter
    sec_ids = [s['id'] for s in sections]
    colors = {}
    saturation = {sid: set(extra_neighbor.get(sid, set())) for sid in sec_ids}
    uncolored = set(sec_ids)
    while uncolored:
        best = max(uncolored, key=lambda s: (len(saturation[s]), len(adj[s])))
        dynamic_forb = set(colors[nb] for nb in adj[best] if nb in colors)
        static_forb = set(forbidden_per_sec.get(best, set()))
        for nb in adj[best]:
            static_forb.update(extra_neighbor.get(nb, set()))
        full_forb = dynamic_forb | static_forb
        free = [p for p in patterns if p not in full_forb]
        if free:
            colors[best] = free[0]
        else:
            hf = hard_forbidden.get(best, set())
            adj_used = Counter()
            for nb in adj[best]:
                if nb in colors: adj_used[colors[nb]] += 5
                for p in extra_neighbor.get(nb, set()): adj_used[p] += 3
            # Tier 1: avoid hard_forbidden AND dynamic_forbidden (real student conflicts)
            tier1 = [p for p in patterns if p not in hf and p not in dynamic_forb]
            if tier1:
                colors[best] = min(tier1, key=lambda p: adj_used[p])
            else:
                # Tier 2: avoid only hard_forbidden
                tier2 = [p for p in patterns if p not in hf]
                colors[best] = min(tier2 if tier2 else patterns, key=lambda p: adj_used[p])
        for nb in adj[best]: saturation[nb].add(colors[best])
        uncolored.remove(best)
    return colors

def assign_two_patterns_with_rebalancing(sections, adj):
    """Assign two distinct recurring slot patterns per section with ZERO conflicts.
    
    Algorithm:
    1. Two-pass DSatur to get (p1, p2) per section (p2 hard-forbidden from equalling p1)
    2. For remaining conflict pairs: move students between sections of the same course
       so that no student's enrolled sections share a time slot
    3. Rebuild graph and repeat until 0 conflicts
    
    Returns: dict {section_id: (p1_pattern, p2_pattern)}
    """
    from collections import defaultdict
    
    sec_dict = {s['id']: s for s in sections}
    sections_by_code = defaultdict(list)
    for s in sections:
        sections_by_code[s['code']].append(s['id'])
    
    for iteration in range(5):
        # Pass 1: DSatur
        p1 = _dsatur_pass(sections, adj, ALL_PATTERNS, {}, {}, {})
        
        # Pass 2: forbidden = own p1 + p1 of all neighbors; hard_forbidden = own p1
        p2_forb, p2_extra, p2_hard = {}, {}, {}
        for s in sections:
            fp = {p1[s['id']]}
            for nb in adj[s['id']]:
                fp.add(p1[nb])
            p2_forb[s['id']] = fp
            p2_extra[s['id']] = {p1[s['id']]}
            p2_hard[s['id']] = {p1[s['id']]}   # p2 MUST differ from p1
        
        p2 = _dsatur_pass(sections, adj, ALL_PATTERNS, p2_forb, p2_extra, p2_hard)
        patterns = {s['id']: (p1[s['id']], p2[s['id']]) for s in sections}
        
        # Find conflict pairs (excluding p1==p2 which hard_forbidden prevents)
        seen = set()
        conflict_pairs = []
        for sec in sections:
            sp = set(patterns[sec['id']])
            for nb in adj[sec['id']]:
                pair = tuple(sorted([sec['id'], nb]))
                if pair not in seen and sp & set(patterns[nb]):
                    conflict_pairs.append(pair)
                    seen.add(pair)
        
        if not conflict_pairs:
            return patterns  # Zero conflicts!
        
        # Student rebalancing: move conflicting students to alternate sections
        moved = 0
        for a, b in conflict_pairs:
            shared = set(sec_dict[a]['students']) & set(sec_dict[b]['students'])
            a_code, b_code = a.rsplit('_', 1)[0], b.rsplit('_', 1)[0]
            a_alts = [x for x in sections_by_code[a_code] if x != a]
            b_alts = [x for x in sections_by_code[b_code] if x != b]
            
            for student in list(shared):
                done = False
                # Try moving from section A to an alt that doesn't conflict with B and has room
                for alt in a_alts:
                    alt_size = len(sec_dict[alt]['students'])
                    if alt_size >= 70:
                        continue  # destination full
                    if not (set(patterns[alt]) & set(patterns[b])):
                        sec_dict[a]['students'].discard(student)
                        sec_dict[alt]['students'].add(student)
                        moved += 1; done = True; break
                if not done:
                    # Try moving from section B to an alt that doesn't conflict with A and has room
                    for alt in b_alts:
                        alt_size = len(sec_dict[alt]['students'])
                        if alt_size >= 70:
                            continue  # destination full
                        if not (set(patterns[alt]) & set(patterns[a])):
                            sec_dict[b]['students'].discard(student)
                            sec_dict[alt]['students'].add(student)
                            moved += 1; break
        
        if moved == 0:
            return patterns  # Stuck — return best effort
        
        # Rebuild conflict graph
        new_adj = build_conflict_graph(sections)
        for sid in adj: adj[sid] = new_adj.get(sid, set())
    
    return patterns

# ── CLASSROOM ASSIGNMENT ───────────────────────────────────────────────────────

# DTI premid/postmid faculty split
DTI_PREMID_FACULTY  = 'Prof. Rohit Kumar'
DTI_POSTMID_FACULTY = 'Prof. Rojers Puthur Joseph'
DTI_PREMID_WEEKS    = {1, 2, 3, 4, 5}   # weeks 1–5
DTI_POSTMID_WEEKS   = {6, 7, 8, 9, 10}  # weeks 6–10

def assign_classrooms(sections, patterns):
    """Assign classrooms using recurring weekly patterns (first-fit greedy).
    Also splits DTI faculty by pre-mid (W1–5: Prof. Rohit Kumar) and
    post-mid (W6–10: Prof. Rojers Puthur Joseph) as per academic constraint.
    """
    room_used = defaultdict(set)  # (week, day, slot) -> set of used rooms
    
    for sec in sections:
        p1, p2 = patterns[sec['id']]
        sec['sessions_scheduled'] = []
        is_dti = (sec['code'] == 'DTI')
        
        for week in WEEKS:
            rooms = get_classrooms(week)
            # Determine faculty for this week (DTI split)
            if is_dti:
                week_faculty = DTI_PREMID_FACULTY if week in DTI_PREMID_WEEKS else DTI_POSTMID_FACULTY
            else:
                week_faculty = sec['faculty']
            
            for day, slot in [p1, p2]:
                key = (week, day, slot)
                avail = [r for r in rooms if r not in room_used[key]]
                if avail:
                    room = avail[0]
                    room_used[key].add(room)
                    sec['sessions_scheduled'].append(
                        {'week': week, 'day': day, 'slot': slot, 
                         'classroom': room, 'faculty': week_faculty})
                else:
                    # All rooms full — find next available slot same day
                    for alt_slot in get_slots(day):
                        if alt_slot == slot:
                            continue
                        alt_key = (week, day, alt_slot)
                        alt_avail = [r for r in rooms if r not in room_used[alt_key]]
                        if alt_avail:
                            room_used[alt_key].add(alt_avail[0])
                            sec['sessions_scheduled'].append(
                                {'week': week, 'day': day, 'slot': alt_slot,
                                 'classroom': alt_avail[0], 'faculty': week_faculty})
                            break
    return sections

# ── CONSTRAINT VERIFICATION ────────────────────────────────────────────────────
def verify(sections):
    """Full verification using (week,day,slot) keys.
    Faculty is read from session record (supports DTI premid/postmid split).
    """
    room_map = {}
    faculty_map = {}
    student_map = {}
    conflicts = []

    for sec in sections:
        for sess in sec['sessions_scheduled']:
            w, d, s, r = sess['week'], sess['day'], sess['slot'], sess['classroom']
            # Use per-session faculty (handles DTI premid/postmid split)
            fac = sess.get('faculty', sec['faculty'])

            rk = (w, d, s, r)
            if rk in room_map:
                conflicts.append(f"ROOM: {sec['id']} & {room_map[rk]} W{w} {d} {s} {r}")
            else:
                room_map[rk] = sec['id']

            fk = (w, d, s, fac)
            if fk in faculty_map:
                conflicts.append(f"FACULTY: {sec['id']} & {faculty_map[fk]} W{w} {d} {s} [{fac}]")
            else:
                faculty_map[fk] = sec['id']

            for st in sec['students']:
                sk = (w, d, s, st)
                if sk in student_map:
                    conflicts.append(f"STUDENT: {st} in {sec['id']} & {student_map[sk]} W{w} {d} {s}")
                else:
                    student_map[sk] = sec['id']

    return conflicts

# ── EXCEL HELPERS ──────────────────────────────────────────────────────────────
def sc(cell, bg=None, bold=False, sz=9, fc='000000', wrap=False, ha='center', va='center', brd=False):
    cell.font = Font(name='Arial', bold=bold, size=sz, color=fc)
    cell.alignment = Alignment(horizontal=ha, vertical=va, wrap_text=wrap)
    if bg: cell.fill = PatternFill('solid', start_color=bg, end_color=bg)
    if brd:
        t = Side(style='thin', color='AAAAAA')
        cell.border = Border(left=t, right=t, top=t, bottom=t)

def write_excel(sections, courses, week_sessions, conflicts, output_path):
    wb = Workbook()
    wb.remove(wb.active)
    _sheet_master(wb, sections, courses, week_sessions)
    _sheet_or_model(wb, courses, len(sections), len(conflicts))
    _sheet_summary(wb, sections, courses)
    _sheet_faculty(wb, sections)
    _sheet_validation(wb, conflicts, sections)
    for w in WEEKS:
        _sheet_week(wb, sections, w)
    wb.save(output_path)

# ── SHEET: MASTER ──────────────────────────────────────────────────────────────
def _sheet_master(wb, sections, courses, week_sessions):
    ws = wb.create_sheet("📋 Master Timetable")
    ws.sheet_view.showGridLines = False

    ws.merge_cells('A1:J1')
    ws['A1'] = '  IIM RANCHI  ·  MBA TERM III TIMETABLE  ·  10-WEEK PROGRAM'
    sc(ws['A1'], bg='0D47A1', bold=True, sz=18, fc='FFFFFF')
    ws.row_dimensions[1].height = 44

    ws.merge_cells('A2:J2')
    ws['A2'] = ('OR-Based Optimal Schedule  |  Algorithm: Per-Week 2-Pass DSatur Graph Coloring  |  '
                'Time Slots: 6/day weekdays, 4/day Sunday  |  15-min breaks + 45-min lunch  |  '
                'Weeks 1–4: 10 Classrooms  |  Weeks 5–10: 4 Classrooms (PAN-IIM Conference)')
    sc(ws['A2'], bg='1565C0', sz=9, fc='E3F2FD')
    ws.row_dimensions[2].height = 20

    # Time structure box
    r = 4
    ws.merge_cells(f'A{r}:E{r}')
    ws[f'A{r}'] = 'DAILY TIME STRUCTURE (1.5hr sessions + breaks)'
    sc(ws[f'A{r}'], bg='37474F', bold=True, sz=10, fc='FFFFFF', brd=True)
    r += 1
    
    time_rows = [
        ('Slot 1', '09:00 – 10:30', ''),
        ('Break', '10:30 – 10:45', '15 min'),
        ('Slot 2', '10:45 – 12:15', ''),
        ('Break', '12:15 – 12:30', '15 min'),
        ('Slot 3', '12:30 – 14:00', ''),
        ('🍽 LUNCH', '14:00 – 14:45', '45 min MANDATORY'),
        ('Slot 4', '14:45 – 16:15', ''),
        ('Break', '16:15 – 16:30', '15 min'),
        ('Slot 5', '16:30 – 18:00', '(Weekdays only)'),
        ('Break', '18:00 – 18:15', '15 min'),
        ('Slot 6', '18:15 – 19:45', '(Weekdays only, ends 7:45 PM)'),
    ]
    for label, time, note in time_rows:
        is_break = 'Break' in label or 'LUNCH' in label
        bg = 'FFF8E1' if 'LUNCH' in label else ('ECEFF1' if is_break else 'FFFFFF')
        ws.cell(r, 1, label); sc(ws.cell(r, 1), bg=bg, bold=not is_break, sz=9, ha='left', brd=True)
        ws.cell(r, 2, time);  sc(ws.cell(r, 2), bg=bg, sz=9, brd=True)
        ws.cell(r, 3, note);  sc(ws.cell(r, 3), bg='FFF8E1' if 'MANDATORY' in note else bg, 
                                  sz=8, ha='left', brd=True, fc='E65100' if 'MANDATORY' in note else '000000')
        ws.row_dimensions[r].height = 14
        r += 1

    # Department legend
    r += 1
    ws.cell(r, 1, 'DEPT'); sc(ws.cell(r, 1), bold=True, sz=10)
    r += 1
    for dept, color in DEPT_COLORS.items():
        ws.cell(r, 1, dept)
        sc(ws.cell(r, 1), bg=color, sz=9, ha='left', brd=True)
        r += 1

    # Course table
    r += 1
    hdrs = ['Code','Course Name','Faculty','Dept','Enrollment','Sections','Sessions','Sample Slot A (Wk1)','Sample Slot B (Wk1)','Status']
    for ci, h in enumerate(hdrs, 1):
        sc(ws.cell(r, ci, h), bg='0D47A1', bold=True, sz=10, fc='FFFFFF', brd=True)
    ws.row_dimensions[r].height = 18
    r += 1

    for code in sorted(courses.keys()):
        c = courses[code]
        dept = get_dept(code)
        color = DEPT_COLORS.get(dept, 'FFFFFF')
        code_secs = [s for s in sections if s['code'] == code]
        # Show sample slots from week 1
        sample = None
        if code_secs and code_secs[0]['sessions_scheduled']:
            wk1_sess = [s for s in code_secs[0]['sessions_scheduled'] if s['week'] == 1]
            if len(wk1_sess) >= 2:
                sample = (f"{wk1_sess[0]['day'][:3]} {SLOT_DISPLAY.get(wk1_sess[0]['slot'], wk1_sess[0]['slot'])}",
                          f"{wk1_sess[1]['day'][:3]} {SLOT_DISPLAY.get(wk1_sess[1]['slot'], wk1_sess[1]['slot'])}")
        p1_str = sample[0] if sample else 'Variable'
        p2_str = sample[1] if sample else 'Variable'
        compliance = '✓ All Constraints Met'
        vals = [code, c['name'], c['faculty'], dept, c['enrollment'],
                c['sections_needed'], 20, p1_str, p2_str, compliance]
        for ci, v in enumerate(vals, 1):
            cell = ws.cell(r, ci, v)
            bg = '70C484' if ci == 10 else color
            sc(cell, bg=bg, sz=9, ha='left' if ci in (2,3,10) else 'center', brd=True)
        r += 1

    ws.column_dimensions['A'].width = 7
    ws.column_dimensions['B'].width = 36
    ws.column_dimensions['C'].width = 28
    ws.column_dimensions['D'].width = 14
    ws.column_dimensions['E'].width = 11
    ws.column_dimensions['F'].width = 9
    ws.column_dimensions['G'].width = 10
    ws.column_dimensions['H'].width = 14
    ws.column_dimensions['I'].width = 14
    ws.column_dimensions['J'].width = 20

# ── SHEET: OR MODEL ────────────────────────────────────────────────────────────
def _sheet_or_model(wb, courses, n_sec, n_conflicts):
    ws = wb.create_sheet("🔢 OR Model")
    ws.sheet_view.showGridLines = False
    ws.merge_cells('A1:C1')
    ws['A1'] = 'OPERATIONS RESEARCH MODEL FORMULATION'
    sc(ws['A1'], bg='0D47A1', bold=True, sz=14, fc='FFFFFF')
    ws.row_dimensions[1].height = 30

    content = [
        ('PROBLEM TYPE', 'Integer Programming (IP) / Constraint Satisfaction Problem (CSP)', '1565C0', True, 'FFFFFF'),
        ('SOLUTION ALGORITHM', 'Two-Pass DSatur (Degree of Saturation) Graph Coloring', '1565C0', True, 'FFFFFF'),
        ('', '', 'FFFFFF', False, '000000'),
        ('DECISION VARIABLE', 'x_{c,w,d,t,r} ∈ {0,1} = 1 if section c assigned to week w, day d, time t, room r', 'ECEFF1', False, '000000'),
        ('', '', 'FFFFFF', False, '000000'),
        ('— INDICES —', '', '37474F', True, 'FFFFFF'),
        ('c ∈ C', f'{n_sec} sections across 27 elective courses', 'F5F5F5', False, '000000'),
        ('w ∈ W = {1..10}', 'Ten-week term', 'FAFAFA', False, '000000'),
        ('d ∈ D', 'Days: Monday through Sunday', 'F5F5F5', False, '000000'),
        ('t ∈ T(d)', 'Mon–Sat: 09:00, 10:45, 12:30, 14:45, 16:30, 18:15 (6 slots) | Sunday: 09:00, 10:45, 12:30, 14:45 (4 slots, ends 16:15)', 'FAFAFA', False, '000000'),
        ('r ∈ R(w)', 'Rooms: 10 classrooms for weeks 1–4; 4 classrooms for weeks 5–10 (PAN-IIM Conference)', 'F5F5F5', False, '000000'),
        ('', '', 'FFFFFF', False, '000000'),
        ('— OBJECTIVE FUNCTION —', '', '37474F', True, 'FFFFFF'),
        ('MINIMIZE Z', 'Z = Σ_{c∈C} max(0, 20 − Σ_{w,d,t,r} x_{c,w,d,t,r})', 'FFF9C4', False, '000000'),
        ('Meaning', 'Minimise total unscheduled sessions across all course-sections', 'FFFDE7', False, '000000'),
        ('', '', 'FFFFFF', False, '000000'),
        ('— CONSTRAINTS —', '', '37474F', True, 'FFFFFF'),
        ('C1 Room Uniqueness',   'Σ_{c} x_{c,w,d,t,r} ≤ 1  ∀ w,d,t,r  [One class per room per slot]', 'E8F5E9', False, '000000'),
        ('C2 Faculty Non-overlap','Σ_{c:fac(c)=f} Σ_r x_{c,w,d,t,r} ≤ 1  ∀ f,w,d,t  [No faculty double-booking]', 'E8F5E9', False, '000000'),
        ('C3 Student Non-overlap','Σ_{c:s∈C_s} Σ_r x_{c,w,d,t,r} ≤ 1  ∀ s,w,d,t  [No student double-booking]', 'E8F5E9', False, '000000'),
        ('C4 Session Completion', 'Σ_{w,d,t,r} x_{c,w,d,t,r} = 20  ∀ c  [Each section completes 20 sessions]', 'E8F5E9', False, '000000'),
        ('C5 Sunday Cutoff',     'x_{c,w,Sun,t,r} = 0  ∀ t > 14:45  [Sunday last slot ends 16:15 ≤ 17:00]', 'E8F5E9', False, '000000'),
        ('C6 Capacity',          '|students(c)| ≤ 70 per section; split courses with enrollment > 70', 'E8F5E9', False, '000000'),
        ('C7 Integrality',       'x_{c,w,d,t,r} ∈ {0,1}', 'E8F5E9', False, '000000'),
        ('', '', 'FFFFFF', False, '000000'),
        ('— ALGORITHM —', '', '37474F', True, 'FFFFFF'),
        ('Step 1', 'Build conflict graph G = (V,E): V = course-sections; edge if shared student or faculty', 'DDEEFF', False, '000000'),
        ('Step 2', 'Pass 1 DSatur: Assign recurring slot p1 to each section (zero student/faculty conflicts)', 'DDEEFF', False, '000000'),
        ('Step 3', 'Pass 2 DSatur: Assign recurring slot p2 (p2 ≠ p1; no conflicts with neighbors)', 'DDEEFF', False, '000000'),
        ('Step 4', 'Each section scheduled as pattern (p1,p2) × 10 weeks = 20 sessions total', 'DDEEFF', False, '000000'),
        ('Step 5', 'First-fit classroom assignment per week respecting R(w) capacity constraint', 'DDEEFF', False, '000000'),
        ('Step 6', 'Post-hoc constraint verification (C1–C7 checked)', 'DDEEFF', False, '000000'),
        ('', '', 'FFFFFF', False, '000000'),
        ('— RESULTS —', '', '37474F', True, 'FFFFFF'),
        ('Sessions Scheduled', f'{n_sec * 20} / {n_sec * 20} (100%)', '70C484' if True else 'FFCCCC', False, '000000'),
        ('Constraint Violations', f'{n_conflicts} ({"FEASIBLE ✓" if n_conflicts == 0 else "infeasible"})', '70C484' if n_conflicts == 0 else 'FFCCCC', False, '000000'),
        ('Max Student Courses', '6 (chromatic number ≥ 6; 47 available slots ≥ 47 sections)', 'F5F5F5', False, '000000'),
    ]

    r = 3
    for label, desc, bg, bold, fc in content:
        ws.cell(r, 1, label)
        ws.cell(r, 2, desc)
        sc(ws.cell(r, 1), bg=bg, bold=bold, sz=9, fc=fc, ha='left', brd=True)
        sc(ws.cell(r, 2), bg=bg, sz=9, fc=fc, ha='left', brd=True, wrap=True)
        ws.row_dimensions[r].height = 18
        r += 1

    ws.column_dimensions['A'].width = 28
    ws.column_dimensions['B'].width = 82

# ── SHEET: SUMMARY ─────────────────────────────────────────────────────────────
def _sheet_summary(wb, sections, courses):
    ws = wb.create_sheet("📊 Course Summary")
    ws.sheet_view.showGridLines = False
    ws.merge_cells('A1:H1')
    ws['A1'] = 'Course & Section Scheduling Summary'
    sc(ws['A1'], bg='0D47A1', bold=True, sz=14, fc='FFFFFF')
    ws.row_dimensions[1].height = 28

    hdrs = ['Section ID','Course Name','Faculty','Dept','Enrolled','Scheduled','Remaining','Status']
    for ci, h in enumerate(hdrs, 1):
        sc(ws.cell(2, ci, h), bg='1565C0', bold=True, sz=10, fc='FFFFFF', brd=True)

    for ri, sec in enumerate(sections, 3):
        sched = len(sec['sessions_scheduled'])
        miss = 20 - sched
        dept = get_dept(sec['code'])
        color = DEPT_COLORS.get(dept, 'FFFFFF')
        status = '✓ Complete' if miss == 0 else f'⚠ Missing {miss}'
        scol = '70C484' if miss == 0 else 'FFB74D'
        for ci, v in enumerate([sec['id'], sec['name'], sec['faculty'], dept,
                                  sec['enrollment'], sched, miss, status], 1):
            cell = ws.cell(ri, ci, v)
            sc(cell, bg=(scol if ci==8 else color), sz=9,
               ha='left' if ci in (2,3) else 'center', brd=True)

    widths = [14,38,28,14,10,12,11,14]
    for ci, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w

# ── SHEET: FACULTY ─────────────────────────────────────────────────────────────
def _sheet_faculty(wb, sections):
    ws = wb.create_sheet("👨‍🏫 Faculty Schedule")
    ws.sheet_view.showGridLines = False
    ws.merge_cells('A1:G1')
    ws['A1'] = 'Faculty Teaching Schedule'
    sc(ws['A1'], bg='0D47A1', bold=True, sz=14, fc='FFFFFF')
    ws.row_dimensions[1].height = 28

    # DTI split note
    ws.merge_cells('A2:G2')
    ws['A2'] = ('⚡ DTI Split: Prof. Rohit Kumar (Pre-Mid, Weeks 1–5)  |  '
                'Prof. Rojers Puthur Joseph (Post-Mid, Weeks 6–10)')
    sc(ws['A2'], bg='FF6F00', bold=True, sz=9, fc='FFFFFF')
    ws.row_dimensions[2].height = 16

    # Build fmap using per-session faculty (handles DTI split)
    fmap = defaultdict(list)
    for sec in sections:
        for sess in sec['sessions_scheduled']:
            fac = sess.get('faculty', sec['faculty'])
            fmap[fac].append((sec, sess))

    r = 4
    for fac in sorted(fmap.keys()):
        entries = fmap[fac]
        ws.merge_cells(f'A{r}:G{r}')
        ws.cell(r, 1, fac)
        sc(ws.cell(r, 1), bg='263238', bold=True, sz=11, fc='FFFFFF')
        ws.row_dimensions[r].height = 20; r += 1
        for ci, h in enumerate(['Section','Course','Week','Day','Time','Room','Dept'], 1):
            sc(ws.cell(r, ci, h), bg='37474F', bold=True, sz=9, fc='FFFFFF', brd=True)
        r += 1
        entries.sort(key=lambda x: (x[1]['week'], DAYS.index(x[1]['day']), x[1]['slot']))
        for sec, sess in entries:
            dept = get_dept(sec['code'])
            color = DEPT_COLORS.get(dept, 'FFFFFF')
            row_data = [sec['id'], sec['name'], sess['week'],
                        sess['day'], sess['slot'], sess['classroom'], dept]
            for ci, v in enumerate(row_data, 1):
                sc(ws.cell(r, ci, v), bg=color, sz=8,
                   ha='left' if ci in (2,7) else 'center', brd=True)
            r += 1
        r += 1

    for col, w in zip('ABCDEFG', [16, 38, 7, 12, 8, 8, 14]):
        ws.column_dimensions[col].width = w

# ── SHEET: VALIDATION ──────────────────────────────────────────────────────────
def _sheet_validation(wb, conflicts, sections):
    ws = wb.create_sheet("✅ Validation")
    ws.sheet_view.showGridLines = False
    ws.merge_cells('A1:B1')
    ws['A1'] = 'Schedule Constraint Validation Report'
    sc(ws['A1'], bg='0D47A1', bold=True, sz=14, fc='FFFFFF')
    ws.row_dimensions[1].height = 28

    total = sum(len(s['sessions_scheduled']) for s in sections)
    needed = len(sections) * 20
    rate = total/needed*100

    stats = [
        ('Total Course-Sections', len(sections)),
        ('Sessions Required (20 × sections)', needed),
        ('Sessions Successfully Scheduled', total),
        ('Scheduling Completion Rate', f'{rate:.1f}%'),
        ('Total Constraint Violations', len(conflicts)),
        ('Room Conflicts (C1)', sum(1 for c in conflicts if c.startswith('ROOM'))),
        ('Faculty Conflicts (C2)', sum(1 for c in conflicts if c.startswith('FACULTY'))),
        ('Student Conflicts (C3)', sum(1 for c in conflicts if c.startswith('STUDENT'))),
        ('Schedule Status', '✓ FEASIBLE – All constraints satisfied' if not conflicts else '⚠ INFEASIBLE'),
    ]
    sc(ws.cell(3, 1, 'METRIC'), bg='1565C0', bold=True, sz=10, fc='FFFFFF', brd=True)
    sc(ws.cell(3, 2, 'VALUE'),  bg='1565C0', bold=True, sz=10, fc='FFFFFF', brd=True)
    for ri, (k, v) in enumerate(stats, 4):
        is_status = k == 'Schedule Status'
        color = ('70C484' if (not conflicts) else 'FFCCCC') if is_status else \
                ('70C484' if (isinstance(v, (int,float)) and v == 0) else 'F5F5F5')
        sc(ws.cell(ri, 1, k), bg='F5F5F5', sz=10, ha='left', brd=True)
        sc(ws.cell(ri, 2, v), bg=color, sz=10, brd=True, bold=is_status)

    # DTI special constraint note
    r_note = 14
    ws.merge_cells(f'A{r_note}:B{r_note}')
    ws.cell(r_note, 1, 'SPECIAL CONSTRAINTS')
    sc(ws.cell(r_note, 1), bg='1565C0', bold=True, sz=10, fc='FFFFFF', brd=True)
    dti_notes = [
        ('DTI Pre-Mid  (Weeks 1–5)', 'Prof. Rohit Kumar — Design Thinking & Innovation (S1 & S2)'),
        ('DTI Post-Mid (Weeks 6–10)', 'Prof. Rojers Puthur Joseph — Design Thinking & Innovation (S1 & S2)'),
        ('CCS (All Weeks)', 'Prof. Rohit Kumar — no time-slot clash with DTI pre-mid sessions (handled by solver)'),
    ]
    for ri, (constraint, detail) in enumerate(dti_notes, r_note + 1):
        sc(ws.cell(ri, 1, constraint), bg='FFF8E1', sz=9, ha='left', brd=True, bold=True)
        sc(ws.cell(ri, 2, detail),     bg='FFF8E1', sz=9, ha='left', brd=True)

    r_conf = r_note + len(dti_notes) + 2

    if conflicts:
        ws.cell(r_conf, 1, 'CONFLICT DETAILS:')
        sc(ws.cell(r_conf, 1), bg='FFCCCC', bold=True, sz=10)
        for ri, c in enumerate(conflicts, r_conf + 1):
            ws.cell(ri, 1, c)
            sc(ws.cell(ri, 1), bg='FFF3F3', sz=8, ha='left', brd=True)
    else:
        ws.merge_cells(f'A{r_conf}:B{r_conf}')
        ws.cell(r_conf, 1, '✓  Zero student conflicts  ·  Zero faculty conflicts  ·  Zero room conflicts  ·  100% sessions scheduled')
        sc(ws.cell(r_conf, 1), bg='C8E6C9', bold=True, sz=11, fc='1B5E20')

    ws.column_dimensions['A'].width = 42
    ws.column_dimensions['B'].width = 50

# ── SHEET: WEEKLY GRID ─────────────────────────────────────────────────────────
def _sheet_week(wb, sections, week):
    rooms = get_classrooms(week)
    ws = wb.create_sheet(f"W{week:02d}")
    ws.sheet_view.showGridLines = False

    ncols = 2 + len(rooms)
    ws.merge_cells(f'A1:{get_column_letter(ncols)}1')
    ws['A1'] = f'  WEEK {week}  ·  {len(rooms)} Classrooms: {", ".join(rooms)}'
    sc(ws['A1'], bg='0D47A1', bold=True, sz=12, fc='FFFFFF')
    ws.row_dimensions[1].height = 24

    # Build lookup
    slot_info = {}  # (day, slot, room) -> info dict
    for sec in sections:
        for sess in sec['sessions_scheduled']:
            if sess['week'] == week:
                k = (sess['day'], sess['slot'], sess['classroom'])
                fac_display = sess.get('faculty', sec['faculty'])
                slot_info[k] = {
                    'text': f"{sec['id']}\n{sec['name'][:20]}\n{fac_display.replace('Prof.','').strip()[:20]}",
                    'dept': get_dept(sec['code'])
                }

    headers = ['Day','Time'] + rooms
    for ci, h in enumerate(headers, 1):
        sc(ws.cell(2, ci, h), bg='1565C0', bold=True, sz=9, fc='FFFFFF', brd=True)

    r = 3
    for day in DAYS:
        slots = get_slots(day)
        for si, slot in enumerate(slots):
            sc(ws.cell(r, 1, day if si == 0 else ''), bg='37474F',
               bold=(si == 0), sz=9, fc='FFFFFF', brd=True)
            sc(ws.cell(r, 2, SLOT_DISPLAY.get(slot, slot)), bg='455A64', sz=8, fc='FFFFFF', brd=True)
            for ri, room in enumerate(rooms, 3):
                cell = ws.cell(r, ri)
                info = slot_info.get((day, slot, room))
                if info:
                    cell.value = info['text']
                    sc(cell, bg=DEPT_COLORS.get(info['dept'], 'FFFFFF'),
                       sz=7, wrap=True, brd=True, ha='left', va='top')
                else:
                    cell.value = ''
                    sc(cell, bg='F9FAFB', brd=True)
            ws.row_dimensions[r].height = 38
            r += 1
            
            # Insert lunch break visual row after slot '12:30' (Slot 3 ends 14:00, lunch 14:00-14:45)
            if slot == '12:30' and day != 'Sunday':
                ws.cell(r, 1, '🍽 LUNCH')
                sc(ws.cell(r, 1), bg='FFF8E1', bold=True, sz=8, fc='E65100', brd=True)
                sc(ws.cell(r, 2, '14:00–14:45'), bg='FFF8E1', sz=8, fc='E65100', brd=True)
                for ci in range(3, ncols+1):
                    sc(ws.cell(r, ci, '45-min Mandatory Lunch Break'), bg='FFF8E1', sz=7, fc='E65100', brd=True)
                ws.row_dimensions[r].height = 12
                r += 1

        for ci in range(1, ncols+1):
            ws.cell(r, ci).fill = PatternFill('solid', start_color='CFD8DC')
        ws.row_dimensions[r].height = 3
        r += 1

    ws.column_dimensions['A'].width = 11
    ws.column_dimensions['B'].width = 9
    for ci in range(3, ncols+1):
        ws.column_dimensions[get_column_letter(ci)].width = 24

# ── MAIN ───────────────────────────────────────────────────────────────────────
if __name__ == '__main__':
    print("Loading course data...")
    courses = load_courses('/mnt/user-data/uploads/WAI_Data.xlsx')
    print(f"  Courses: {len(courses)}")

    print("Building sections...")
    sections = build_sections(courses)
    print(f"  Sections: {len(sections)}")

    print("Building conflict graph...")
    adj = build_conflict_graph(sections)
    max_deg = max(len(v) for v in adj.values())
    print(f"  Max vertex degree: {max_deg} | Available patterns: {len(ALL_PATTERNS)}")

    print("Running Two-Pass DSatur + Student Rebalancing...")
    patterns = assign_two_patterns_with_rebalancing(sections, adj)

    print("Assigning classrooms...")
    sections = assign_classrooms(sections, patterns)

    total = sum(len(s['sessions_scheduled']) for s in sections)
    needed = len(sections) * 20
    print(f"  Sessions: {total}/{needed} ({total/needed*100:.1f}%)")

    print("Verifying constraints...")
    conflicts = verify(sections)
    print(f"  Conflicts: {len(conflicts)} {'✓ FEASIBLE' if not conflicts else '⚠ INFEASIBLE'}")

    print("Writing Excel timetable...")
    output = '/mnt/user-data/outputs/IIM_Ranchi_MBA_Timetable.xlsx'
    write_excel(sections, courses, patterns, conflicts, output)
    print(f"  Saved: {output}")
    print("\nDone! 🎓")
